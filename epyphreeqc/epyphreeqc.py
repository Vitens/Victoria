from solver import Solver
from water_quality import Water_quality

import datetime
import openpyxl
import os


class EpyPhreeqc(object):

    def __init__(self, network, sol_dict, pp, time_1):
        self.solver = Solver(network, sol_dict, pp)
        self.quality = Water_quality(pp)
        self.output = []
        self.time = time_1

    def run(self, network, sol_dict, timestep,
            start_time, total_time, reporting_timestep, time_1):
        # Main calculation loop
        self.output = []
        # Check the flow direction in each pipe
        self.solver.check_connections(network)
        counter = start_time / timestep
        date_step = datetime.timedelta(minutes=reporting_timestep)

        for t in range(start_time, total_time, timestep):
            # Timestep in minutes
            self.solver.step(network, timestep, sol_dict)
            print('TIME', t)

            # Store output at reporting timestep
            if t % reporting_timestep == 0:
                self.quality.nodes(network, self.solver.models)
                self.quality.pipes(network, self.solver.models)

                self.output.append(self.quality.get_quality_nodes(network, 'Ca', 'mg'))

                node_list = ['56159097', '56157441', '56158933', '56158415', '56158831']
                species_list = ['Ca']
                unit = 'mg'

                self.time += date_step
                # self.print_output(node_list, species_list, unit, counter, self.time)
                counter += 1

        return self.output

    def initialize_file(self, node_list, species_list, unit):
        try:
            os.remove('data.xlsx')
        finally:

            wb = openpyxl.Workbook()

            for node in node_list:
                wb.create_sheet(node)
                sheet = wb[node]
                sheet.cell(row=1, column=1, value='Time')
                col = 2
                for spec in species_list:
                    sheet.cell(row=1, column=col, value=spec)
                    col += 1

            wb.save('data.xlsx')

    def print_output(self, node_list, species_list, unit, counter, time):
        # print output in file:

        wb = openpyxl.load_workbook('data.xlsx')

        for node in node_list:
            sheet = wb[node]
            col = 2
            row = 3 + counter
            sheet.cell(row=row, column=1, value=time)
            for spec in species_list:
                temp = self.quality.get_quality_node(node, spec, unit)
                sheet.cell(row=row, column=col, value=temp['q'])

                col += 1

        wb.save('data.xlsx')
